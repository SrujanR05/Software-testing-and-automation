from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class ExcelReader:
    def __init__(self, file_path: str | Path) -> None:
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

    def sheet_names(self) -> list[str]:
        workbook = load_workbook(self.file_path, data_only=True)
        return workbook.sheetnames

    def read_rows(self, sheet_name: str | None = None) -> list[dict[str, Any]]:
        workbook = load_workbook(self.file_path, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(header).strip() if header is not None else f"column_{index + 1}" for index, header in enumerate(rows[0])]
        records: list[dict[str, Any]] = []
        for row in rows[1:]:
            record = {headers[index]: value for index, value in enumerate(row)}
            records.append(record)
        return records

    def read_cell(self, sheet_name: str, row: int, column: int) -> Any:
        workbook = load_workbook(self.file_path, data_only=True)
        worksheet = workbook[sheet_name]
        return worksheet.cell(row=row, column=column).value
